"""
BOM数据加载器 - 从Excel解析扫地机器人拆机数据

Excel 格式约定（每个 Sheet 对应一款产品）：
  - PCB/芯片 section：列依次为 PCB板、类别、功能模块、芯片型号、数量、厂商、规格、单价、价格
  - 电机 section：以第0列"电机"/"Motor"为分隔行标记
  - 传感器 section：以第0列"传感器"/"Sensor"为分隔行标记
  - 其他 section：以第0列"其他"/"Other"为分隔行标记

自定义 Excel 路径：
  - 在 data/ 目录放置 Excel 文件，路径通过环境变量 BOM_EXCEL_FILE 指定
  - 或在 data/ 目录放置任意 .xlsx 文件（自动识别第一个）
  - 若未找到 Excel，get_bom_data() 返回空字典，系统正常运行（仅网络调研模式）
"""
from __future__ import annotations

import os
from pathlib import Path
from typing import Any

import openpyxl


# ─── Excel 路径解析（不硬编码文件名）────────────────────────────
def _find_excel() -> Path | None:
    """查找 BOM Excel 文件：优先环境变量，其次 data/ 目录下第一个 .xlsx"""
    env_path = os.getenv("BOM_EXCEL_FILE", "")
    if env_path:
        p = Path(env_path)
        return p if p.exists() else None

    data_dir = Path(__file__).parent.parent / "data"
    xlsx_files = sorted(data_dir.glob("*.xlsx"))
    return xlsx_files[0] if xlsx_files else None


# ─── 原始数据缓存 ───────────────────────────────────────────────
_bom_cache: dict[str, dict[str, Any]] | None = None


def _parse_excel() -> dict[str, dict[str, Any]]:
    """解析 Excel，返回 {型号: {pcb, motors, sensors, others}} 结构"""
    excel_file = _find_excel()
    if excel_file is None:
        return {}
    wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
    result: dict[str, dict[str, Any]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = [row for row in ws.iter_rows(values_only=True) if any(v is not None for v in row)]

        pcb_rows: list[tuple] = []
        motor_rows: list[tuple] = []
        sensor_rows: list[tuple] = []
        other_rows: list[tuple] = []

        current: list[tuple] = pcb_rows
        for row in rows:
            first = str(row[0]).strip() if row[0] else ""
            if first in ("电机", "Motor"):
                current = motor_rows
                continue
            if first in ("传感器", "Sensor"):
                current = sensor_rows
                continue
            if first in ("其他", "Other"):
                current = other_rows
                continue
            current.append(row)

        result[sheet_name] = {
            "pcb":     _parse_pcb(pcb_rows),
            "motors":  _parse_motors(motor_rows),
            "sensors": _parse_sensors(sensor_rows),
            "others":  _parse_others(other_rows),
        }

    wb.close()
    return result


def _parse_pcb(rows: list[tuple]) -> list[dict]:
    if not rows:
        return []
    header = rows[0] if rows else ()
    items: list[dict] = []
    current_board = ""
    current_sub = ""

    # 科沃斯格式有"类别2"列
    has_sub_category = (
        len(header) > 2
        and header[2] is not None
        and "类别2" in str(header[2])
    )

    for row in rows[1:]:
        cols = list(row) + [None] * 12

        if has_sub_category:
            board1 = str(cols[1]).strip().replace("\n", "") if cols[1] else ""
            board2 = str(cols[2]).strip().replace("\n", "") if cols[2] else ""
            func   = str(cols[3]).strip() if cols[3] else ""
            model  = str(cols[4]).strip() if cols[4] else ""
            qty    = cols[5]
            mfr    = str(cols[6]).strip() if cols[6] else ""
            spec   = str(cols[7]).strip() if cols[7] else ""
            price  = cols[8]
            total  = cols[9]
            if board1 and board1 != "None":
                current_board = board1
            if board2 and board2 != "None":
                current_sub = board2
        else:
            board1 = str(cols[1]).strip().replace("\n", "") if cols[1] else ""
            func   = str(cols[2]).strip() if cols[2] else ""
            model  = str(cols[3]).strip() if cols[3] else ""
            qty    = cols[4]
            mfr    = str(cols[5]).strip() if cols[5] else ""
            spec   = str(cols[6]).strip() if cols[6] else ""
            price  = cols[7]
            total  = cols[8]
            board2 = ""
            if board1 and board1 != "None":
                current_board = board1
                current_sub = ""

        func  = func  if func  not in ("", "None") else ""
        model = model if model not in ("", "None") else ""

        if not func and not model:
            continue

        items.append({
            "board":      current_board,
            "sub_board":  current_sub if has_sub_category else "",
            "function":   func,
            "model":      model,
            "qty":        qty,
            "manufacturer": mfr if mfr not in ("", "None") else "",
            "spec":       spec if spec not in ("", "None") else "",
            "unit_price": price,
            "total_price": total,
        })

    return items


def _parse_motors(rows: list[tuple]) -> list[dict]:
    items: list[dict] = []
    for row in rows[1:]:
        cols = list(row) + [None] * 10
        name  = str(cols[1]).strip() if cols[1] else ""
        mtype = str(cols[2]).strip() if cols[2] else ""
        model = str(cols[3]).strip() if cols[3] else ""
        params = str(cols[4]).strip() if cols[4] else ""
        qty   = cols[5]
        mfr   = str(cols[6]).strip() if cols[6] else ""
        if not name or name == "None":
            continue
        items.append({
            "name":         name,
            "type":         mtype  if mtype  != "None" else "",
            "model":        model  if model  != "None" else "",
            "params":       params if params != "None" else "",
            "qty":          qty,
            "manufacturer": mfr    if mfr    != "None" else "",
        })
    return items


def _parse_sensors(rows: list[tuple]) -> list[dict]:
    items: list[dict] = []
    for row in rows[1:]:
        cols  = list(row) + [None] * 10
        name  = str(cols[1]).strip() if cols[1] else ""
        stype = str(cols[2]).strip() if cols[2] else ""
        qty   = cols[3]
        mfr   = str(cols[4]).strip() if cols[4] else ""
        price = cols[5]
        note  = str(cols[6]).strip() if cols[6] else ""
        if not name or name == "None":
            continue
        items.append({
            "name":         name,
            "type":         stype if stype != "None" else "",
            "qty":          qty,
            "manufacturer": mfr   if mfr   != "None" else "",
            "unit_price":   price,
            "note":         note  if note  != "None" else "",
        })
    return items


def _parse_others(rows: list[tuple]) -> list[dict]:
    items: list[dict] = []
    for row in rows[1:]:
        cols  = list(row) + [None] * 10
        name  = str(cols[1]).strip() if cols[1] else ""
        itype = str(cols[2]).strip() if cols[2] else ""
        spec  = str(cols[3]).strip() if cols[3] else ""
        mfr   = str(cols[4]).strip() if cols[4] else ""
        price = cols[5]
        if not name or name == "None":
            continue
        items.append({
            "name":         name,
            "type":         itype if itype != "None" else "",
            "spec":         spec  if spec  != "None" else "",
            "manufacturer": mfr   if mfr   != "None" else "",
            "price":        price,
        })
    return items


def get_bom_data() -> dict[str, dict[str, Any]]:
    """返回所有产品的 BOM 数据。未找到 Excel 文件时返回空字典。"""
    global _bom_cache
    if _bom_cache is None:
        _bom_cache = _parse_excel()
    return _bom_cache


def get_models() -> list[str]:
    return list(get_bom_data().keys())
