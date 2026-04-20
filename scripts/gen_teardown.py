#!/usr/bin/env python3
"""
通用拆机分析 Excel 生成器（含 Claude API 网络数据补全）

用法：
    python scripts/gen_teardown.py "石头G30S Pro"              # 查找/生成CSV → 补全价格 → 输出Excel
    python scripts/gen_teardown.py "科沃斯X8 Pro" --msrp 6999
    python scripts/gen_teardown.py "石头G30S Pro" --no-enrich  # 跳过价格网络补全
    python scripts/gen_teardown.py "石头G30S Pro" --enrich     # 强制重新补全全部价格
    python scripts/gen_teardown.py --csv data/teardowns/xxx.csv "机型名"

输出：data/{model_slug}_拆机分析.xlsx
需要环境变量：ANTHROPIC_API_KEY
"""
from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
from pathlib import Path
from typing import Optional

import anthropic
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT        = Path(__file__).parent.parent
DATA_DIR    = ROOT / "data"
TEARDOWN_DIR = DATA_DIR / "teardowns"

# ── BOM 8桶 ──────────────────────────────────────────────────────
BUCKETS = [
    ("compute_electronics", "计算/电子"),
    ("perception",          "感知"),
    ("power_motion",        "驱动运动"),
    ("cleaning",            "清洁"),
    ("energy",              "能源/电池"),
    ("dock_station",        "基站"),
    ("structure_cmf",       "结构CMF"),
    ("mva_software",        "MVA+软件"),
]
BUCKET_MAP = {k: v for k, v in BUCKETS}

# 旗舰机各桶理论占比区间
BUCKET_THEORY = {
    "compute_electronics": (0.10, 0.12),
    "perception":          (0.10, 0.13),
    "power_motion":        (0.10, 0.12),
    "cleaning":            (0.13, 0.17),
    "energy":              (0.07, 0.09),
    "dock_station":        (0.15, 0.20),
    "structure_cmf":       (0.10, 0.13),
    "mva_software":        (0.09, 0.13),
}

CSV_FIELDS = [
    "bom_bucket", "section", "name", "model", "type",
    "spec", "manufacturer", "unit_price", "qty", "confidence", "product_source",
]

# ── FCC 厂商代码 ───────────────────────────────────────────────
# 格式：https://fccid.io/{grantee_code}  列出该品牌全部已认证设备
# 单设备：https://fccid.io/{grantee_code}{product_code}
BRAND_FCC_CODE: dict[str, str] = {
    "石头":     "2AN2O",   # Roborock
    "roborock": "2AN2O",
    "云鲸":     "2ARZZ",   # Narwal
    "narwal":   "2ARZZ",
    "追觅":     "2AX54",   # Dreame
    "dreame":   "2AX54",
    "科沃斯":   "2A6HE",   # Ecovacs
    "ecovacs":  "2A6HE",
}


def _fcc_hint(model: str) -> str:
    """
    根据机型名称推断 FCC grantee code 并构造搜索提示。
    返回空字符串表示品牌未知。
    """
    low = model.lower()
    for keyword, code in BRAND_FCC_CODE.items():
        if keyword in low:
            return (
                f"FCC grantee code: {code}\n"
                f"- 品牌设备列表: https://fccid.io/{code}\n"
                f"- 在列表中找到型号最相近的设备，点击进入详情页\n"
                f"- 用 web_fetch 抓取该设备的 Internal Photos（内部照片）和 "
                f"Block Diagram（框图/原理图）页面，从照片中识别 PCB 上的芯片型号、\n"
                f"  主板布局和主要元器件，从框图中提取系统架构和子系统划分"
            )
    return ""

# ── Excel 样式 ─────────────────────────────────────────────────
CONF = {
    "teardown": dict(bg="C6EFCE", fg="375623", label="实物拆机"),
    "web":      dict(bg="BDD7EE", fg="1F4E79", label="网络调研"),
    "estimate": dict(bg="FFEB9C", fg="9C5700", label="行业估算"),
}
HDR_FILL  = PatternFill("solid", fgColor="2F5496")
HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
GRAY_FILL = PatternFill("solid", fgColor="F2F2F2")
GRAY_FONT = Font(size=10, color="595959", italic=True)
_THIN     = Side(style="thin", color="BFBFBF")
BORDER    = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
FMT_MONEY = '#,##0.00'
FMT_PCT   = '0.0%'


# ══════════════════════════════════════════════════════════════════
#  工具函数
# ══════════════════════════════════════════════════════════════════

def _slug(model: str) -> str:
    """'石头 G30S Pro' → '石头G30SPro'"""
    return re.sub(r"[\s\-]+", "", model)


def _fill(bg: str) -> PatternFill:
    return PatternFill("solid", fgColor=bg)


def _c(ws, row: int, col: int, value=None, *,
       bold=False, bg=None, fg="000000",
       align="left", wrap=False, fmt=None,
       italic=False, border=True):
    c = ws.cell(row, col, value)
    c.font = Font(bold=bold, color=fg, size=10, italic=italic)
    if bg:
        c.fill = _fill(bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if fmt:
        c.number_format = fmt
    if border:
        c.border = BORDER
    return c


# ══════════════════════════════════════════════════════════════════
#  CSV 读写
# ══════════════════════════════════════════════════════════════════

def load_csv(path: Path) -> list[dict]:
    with open(path, encoding="utf-8") as f:
        return list(csv.DictReader(f))


def save_csv(rows: list[dict], path: Path, model: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    for r in rows:
        r.setdefault("product_source", model)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=CSV_FIELDS, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def find_csv(model: str) -> Optional[Path]:
    """按 slug 精确/模糊匹配 teardowns/ 下的 CSV 文件"""
    slug = _slug(model)
    exact = TEARDOWN_DIR / f"{slug}_teardown.csv"
    if exact.exists():
        return exact
    for p in sorted(TEARDOWN_DIR.glob("*_teardown*.csv")):
        if slug.lower() in p.name.lower() or p.stem.lower().startswith(slug.lower()[:6]):
            return p
    return None


def _norm_price(val) -> float:
    try:
        return float(val or 0)
    except (ValueError, TypeError):
        return 0.0


def _norm_qty(val) -> int:
    try:
        return max(1, int(val or 1))
    except (ValueError, TypeError):
        return 1


# ══════════════════════════════════════════════════════════════════
#  Claude API 工具调用循环（服务端 web_search / web_fetch）
# ══════════════════════════════════════════════════════════════════

SERVER_TOOLS = [
    {"type": "web_search_20260209", "name": "web_search"},
    {"type": "web_fetch_20260209",  "name": "web_fetch"},
]


def _run_web_agent(system: str, user: str,
                   max_tokens: int = 8192) -> str:
    """
    用 Claude + 服务端 web_search/web_fetch 完成一次问答，返回最终文本。
    - pause_turn：服务端工具尚未结束，继续循环
    - end_turn：收集所有 text block 返回
    """
    client = anthropic.Anthropic()
    messages: list[dict] = [{"role": "user", "content": user}]

    while True:
        resp = client.messages.create(
            max_tokens=max_tokens,
            system=[{
                "type": "text",
                "text": system,
                "cache_control": {"type": "ephemeral"},
            }],
            tools=SERVER_TOOLS,
            messages=messages,
        )

        messages.append({"role": "assistant", "content": resp.content})

        if resp.stop_reason == "end_turn":
            texts = [b.text for b in resp.content if getattr(b, "type", "") == "text"]
            return "\n".join(texts)

        if resp.stop_reason == "pause_turn":
            # 服务端工具仍在执行，直接继续循环
            continue

        # tool_use（客户端工具，此脚本不注册，通常不会到达）
        texts = [b.text for b in resp.content if getattr(b, "type", "") == "text"]
        if texts:
            return "\n".join(texts)
        raise RuntimeError(f"意外的 stop_reason: {resp.stop_reason}")


def _extract_json_array(text: str) -> list[dict]:
    """从 Claude 响应中提取第一个 JSON 数组"""
    # 去除 ```json ... ``` 包裹
    text = re.sub(r"```(?:json)?\s*", "", text).strip().rstrip("`").strip()
    start = text.find("[")
    end   = text.rfind("]") + 1
    if start == -1 or end == 0:
        raise ValueError(f"响应中未找到 JSON 数组（前500字）：\n{text[:500]}")
    return json.loads(text[start:end])


# ══════════════════════════════════════════════════════════════════
#  从网络生成完整拆机清单
# ══════════════════════════════════════════════════════════════════

_GEN_SYSTEM = (
    "你是扫地机器人硬件 BOM 成本分析专家，熟悉各品牌拆机报告、"
    "主要芯片厂商（全志、瑞芯微、比特大陆、TI、ST、InvenSense）及元件市场价。"
    "你善于从 FCC ID.io 的内部照片中识别 PCB 芯片丝印和器件型号。"
    "请严格按照指定 JSON 格式输出，不要输出任何额外文字。"
)

_GEN_PROMPT = """\
请为 **{model}**（建议零售价约 {msrp} 元）生成一份完整的拆机 BOM 清单。

**操作步骤（按顺序执行）**

1. **FCC 认证文件**（优先级最高，可获取内部照片和原理图）
{fcc_hint}
   - 若品牌未知，用 web_search 搜索 "{model} FCC ID" 获取认证编号

2. **拆机报告搜索**
   - web_search: "{model} 拆机报告 PCB 芯片" / "{model} teardown internals"
   - web_search: "{model} disassembly main board chip"

3. **零件价格查询**
   - 对步骤1-2中识别出的关键芯片/模组，web_search 查批量采购价
   - 重点查：雷达模组、SoC/CPU、NPU、结构光/ToF 模组、电池包

4. **综合所有来源**，填写下方 JSON

---

**FCC ID.io 使用技巧**
- 设备列表页：https://fccid.io/{{grantee_code}}
- 进入设备详情后，优先查看：
  - **Internal Photos**（内部照片）→ 从 PCB 丝印识别芯片型号
  - **Block Diagram**（框图）→ 了解系统架构和子系统划分
  - **External Photos**（外观照片）→ 确认产品版本
  - **Test Report**（测试报告）→ 有时包含关键元件清单
- 用 web_fetch 抓取照片页，仔细描述能看到的所有芯片丝印文字

---

**8个BOM桶说明**
- compute_electronics: SoC/CPU、NPU、MCU、RAM/ROM、Wi-Fi/BT、PMIC、马达驱动IC、充电IC、被动元件、PCB板
- perception: 激光雷达、结构光/ToF、IMU、下视/沿墙/碰撞传感器、超声波
- power_motion: 风机、驱动轮电机+齿轮箱+减震、底盘升降电机及机构
- cleaning: 拖布盘/电机、水泵、机身水箱、滚刷/边刷本体、管路密封
- energy: 电池包（电芯+BMS）
- dock_station: 集尘电机、清洗水泵、加热烘干模块、基站PCB、基站外壳/水箱
- structure_cmf: 机身上盖/底盘注塑、保险杠、万向轮、尘盒、喷涂/CMF、模具摊销
- mva_software: 组装人工、SLAM算法版税、包装材料、QA出厂检测、OS/系统授权

---

**输出格式**（直接输出 JSON 数组，不要 markdown 代码块）：
[
  {{
    "bom_bucket": "compute_electronics",
    "section": "PCB",
    "name": "CPU",
    "model": "具体型号（从FCC照片/拆机报告识别，无法确定则填空字符串）",
    "type": "类型描述",
    "spec": "规格参数",
    "manufacturer": "厂商",
    "unit_price": 40,
    "qty": 1,
    "confidence": "teardown"
  }}
]

confidence 说明：
- teardown：从 FCC 内部照片/拆机照片直接识别（最高可信）
- web：从网络评测/规格页确认
- estimate：行业基准估算（无直接证据）

**目标**：总成本约 {bom_target} 元（零售价的 50%），覆盖全部 8 个桶，每桶至少 3 个主要零件。\
"""


def generate_from_web(model: str, msrp: float) -> list[dict]:
    """调用 Claude API + web_search/web_fetch（含 FCC ID.io）从零生成拆机清单"""
    print(f"  → 调用 Claude API 网络调研 {model} 拆机数据…")
    fcc_hint = _fcc_hint(model)
    if fcc_hint:
        print(f"  → 检测到 FCC 代码，将优先抓取 FCC ID.io 内部照片")
    bom_target = int(msrp * 0.50)
    prompt = _GEN_PROMPT.format(
        model=model, msrp=int(msrp), bom_target=bom_target,
        fcc_hint=fcc_hint if fcc_hint else "   （未知品牌，跳过，直接进行步骤2）",
    )

    text = _run_web_agent(_GEN_SYSTEM, prompt, max_tokens=8192)
    rows = _extract_json_array(text)

    for r in rows:
        r["product_source"] = model
        for key in CSV_FIELDS:
            r.setdefault(key, "")
        r["unit_price"] = _norm_price(r.get("unit_price"))
        r["qty"]        = _norm_qty(r.get("qty"))

    print(f"  ✓ 生成 {len(rows)} 条零件记录")
    return rows


# ══════════════════════════════════════════════════════════════════
#  网络补全缺失价格
# ══════════════════════════════════════════════════════════════════

_ENRICH_SYSTEM = (
    "你是扫地机器人元件价格专家，熟悉立创商城、LCSC、Mouser、嘉立创等平台的批量采购价。"
    "如果零件有具体型号，优先在元件平台查询；"
    "如果型号未知，可先从 FCC ID.io 内部照片确认型号再查价。"
    "严格只输出 JSON 数组，不要其他任何文字。"
)

_ENRICH_PROMPT = """\
以下是 **{model}** 拆机清单中价格缺失的零件，请通过 web_search 查询它们的市场价（人民币，\
批量 1000+ pcs）。

```json
{items_json}
```

返回与输入**顺序一致**的 JSON 数组，每项包含：
- name: 保持原样
- unit_price: 查到的批量价（元），查不到则填行业基准估算值
- manufacturer: 厂商（如原本已有则保持）
- confidence: "web"（查到真实价格）或 "estimate"（行业估算）

只输出 JSON 数组。\
"""


def enrich_prices(rows: list[dict], model: str, force: bool = False) -> list[dict]:
    """调用 Claude + web_search 补全缺失的 unit_price / manufacturer"""
    need = [
        (i, r) for i, r in enumerate(rows)
        if force or _norm_price(r.get("unit_price")) == 0
    ]
    if not need:
        print("  → 所有零件已有价格，跳过网络补全")
        return rows

    print(f"  → {len(need)} 个零件价格缺失，调用 Claude API 补全…")

    BATCH = 25
    for b_start in range(0, len(need), BATCH):
        batch = need[b_start : b_start + BATCH]
        items = [
            {
                "name":         r.get("name", ""),
                "model":        r.get("model", ""),
                "type":         r.get("type", ""),
                "spec":         r.get("spec", ""),
                "manufacturer": r.get("manufacturer", ""),
                "bom_bucket":   r.get("bom_bucket", ""),
            }
            for _, r in batch
        ]
        prompt = _ENRICH_PROMPT.format(
            model=model,
            items_json=json.dumps(items, ensure_ascii=False, indent=2),
        )
        try:
            text    = _run_web_agent(_ENRICH_SYSTEM, prompt, max_tokens=4096)
            results = _extract_json_array(text)
        except Exception as e:
            print(f"  ⚠ 批次 {b_start // BATCH + 1} 失败: {e}，跳过")
            continue

        for j, (i, _) in enumerate(batch):
            if j >= len(results):
                break
            res = results[j]
            if _norm_price(res.get("unit_price")):
                rows[i]["unit_price"] = _norm_price(res["unit_price"])
            if res.get("manufacturer") and not rows[i].get("manufacturer"):
                rows[i]["manufacturer"] = res["manufacturer"]
            if res.get("confidence"):
                rows[i]["confidence"] = res["confidence"]

        print(f"  ✓ 批次 {b_start // BATCH + 1} 完成（{len(batch)} 条）")

    return rows


# ══════════════════════════════════════════════════════════════════
#  Excel 生成（Sheet 1: 拆机清单 / Sheet 2: 成本对比）
# ══════════════════════════════════════════════════════════════════

MAIN_COLS = [
    ("BOM分类",      11),
    ("模块",          8),
    ("名称",         14),
    ("型号",         15),
    ("类型",         13),
    ("规格",         24),
    ("厂商",         11),
    ("理论单价/元",   9),
    ("数量",          5),
    ("理论小计/元",   9),
    ("实测价格/元",  11),
    ("实测小计/元",  10),
    ("价格偏差",      9),
    ("数据来源",      9),
]

COMP_COLS = [
    ("BOM分类",      14),
    ("理论成本/元",  13),
    ("理论占比%",    10),
    ("实测成本/元",  13),
    ("实测占比%",    10),
    ("偏差金额/元",  13),
    ("偏差%",         9),
    ("校对备注",     22),
]


def _build_sheet_main(ws, rows: list[dict], model: str) -> tuple[int, int]:
    ws.title = "拆机清单"

    ws.merge_cells("A1:N1")
    t = ws.cell(1, 1, f"{model} · 拆机电控分析（置信度标注）")
    t.font = Font(bold=True, size=13, color="1F3864")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:N2")
    leg = ws.cell(
        2, 1,
        "置信度：  🟢 teardown 实物拆机（最高）  "
        "🔵 web 网络调研  "
        "🟡 estimate 行业估算（最低）  ‖  "
        "K列「实测价格」请工程师依拆机结果填写，L/M 列自动计算",
    )
    leg.font = Font(italic=True, size=9, color="595959")
    leg.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 30
    for ci, (hdr, w) in enumerate(MAIN_COLS, 1):
        c = ws.cell(3, ci, hdr)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w

    DATA_START = 4
    r = DATA_START
    CENTER_COLS = {8, 9, 10, 11, 12, 13}

    for row in rows:
        conf  = (row.get("confidence") or "estimate").strip()
        style = CONF.get(conf, CONF["estimate"])
        bg, fg = style["bg"], style["fg"]

        up  = _norm_price(row.get("unit_price"))
        qty = _norm_qty(row.get("qty"))
        bucket_label = BUCKET_MAP.get(row.get("bom_bucket", ""), row.get("bom_bucket", ""))

        vals = [
            bucket_label,
            row.get("section", ""),
            row.get("name", ""),
            row.get("model", ""),
            row.get("type", ""),
            row.get("spec", ""),
            row.get("manufacturer", ""),
            up or None,
            qty,
        ]
        for ci, v in enumerate(vals, 1):
            _c(ws, r, ci, v,
               bg=bg, fg=fg,
               align="center" if ci in CENTER_COLS else "left",
               wrap=(ci == 6),
               fmt=FMT_MONEY if ci == 8 else None)

        # J 理论小计
        jc = ws.cell(r, 10)
        jc.fill = _fill(bg); jc.border = BORDER
        jc.alignment = Alignment(horizontal="center", vertical="center")
        jc.number_format = FMT_MONEY
        jc.font = Font(color=fg, size=10)
        jc.value = (up * qty) if up else None

        # K 实测价格（工程师填写）
        kc = ws.cell(r, 11)
        kc.fill = GRAY_FILL; kc.border = BORDER
        kc.alignment = Alignment(horizontal="center", vertical="center")
        kc.number_format = FMT_MONEY; kc.font = GRAY_FONT

        # L 实测小计
        lc = ws.cell(r, 12, f'=IF(K{r}="","",K{r}*I{r})')
        lc.fill = GRAY_FILL; lc.border = BORDER
        lc.alignment = Alignment(horizontal="center", vertical="center")
        lc.number_format = FMT_MONEY; lc.font = GRAY_FONT

        # M 偏差
        mc = ws.cell(r, 13, f'=IFERROR((L{r}-J{r})/J{r},"")')
        mc.fill = GRAY_FILL; mc.border = BORDER
        mc.alignment = Alignment(horizontal="center", vertical="center")
        mc.number_format = FMT_PCT; mc.font = GRAY_FONT

        # N 数据来源
        _c(ws, r, 14, style["label"], bg=bg, fg=fg, align="center")

        ws.row_dimensions[r].height = 18
        r += 1

    ws.freeze_panes = "A4"
    return DATA_START, r - 1


def _build_sheet_comparison(ws, rows: list[dict], model: str,
                             msrp: float, ds: int, de: int) -> None:
    ws.title = "成本对比"

    ws.merge_cells("A1:H1")
    t = ws.cell(1, 1, f"{model} · 理论成本 vs 实测成本 — 分类对比")
    t.font = Font(bold=True, size=13, color="1F3864")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:H2")
    note = ws.cell(
        2, 1,
        f"零售价 {int(msrp)} 元，旗舰机 BOM 率约 50%（理论总成本 ≈ {int(msrp*0.5)} 元）。"
        "理论成本 = AI 估算；实测成本 = 工程师填写「拆机清单」K列后自动汇总。",
    )
    note.font = Font(italic=True, size=9, color="595959")
    note.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 30
    for ci, (hdr, w) in enumerate(COMP_COLS, 1):
        c = ws.cell(3, ci, hdr)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w

    # 各桶理论成本
    theory: dict[str, float] = {k: 0.0 for k, _ in BUCKETS}
    for row in rows:
        bkt = row.get("bom_bucket", "").strip()
        total = _norm_price(row.get("unit_price")) * _norm_qty(row.get("qty"))
        if bkt in theory:
            theory[bkt] += total

    total_theory = sum(theory.values())

    r = 4
    for bkt, label in BUCKETS:
        th     = theory.get(bkt, 0.0)
        th_pct = th / total_theory if total_theory else 0.0
        lo, hi = BUCKET_THEORY.get(bkt, (0, 0))

        _c(ws, r, 1, label, bold=True)

        bc = ws.cell(r, 2, th)
        bc.number_format = FMT_MONEY; bc.border = BORDER
        bc.alignment = Alignment(horizontal="center", vertical="center")
        bc.font = Font(size=10)

        cc = ws.cell(r, 3, th_pct)
        cc.number_format = FMT_PCT; cc.border = BORDER
        cc.alignment = Alignment(horizontal="center", vertical="center")
        cc.font = Font(size=10)

        # D 实测成本 SUMIF
        total_r = 4 + len(BUCKETS)
        dc = ws.cell(r, 4,
            f"=SUMIF('拆机清单'!$A${ds}:$A${de},A{r},'拆机清单'!$L${ds}:$L${de})")
        dc.number_format = FMT_MONEY; dc.border = BORDER
        dc.fill = GRAY_FILL; dc.font = GRAY_FONT
        dc.alignment = Alignment(horizontal="center", vertical="center")

        # E 实测占比
        ec = ws.cell(r, 5, f'=IFERROR(D{r}/$D${total_r},"")')
        ec.number_format = FMT_PCT; ec.border = BORDER
        ec.fill = GRAY_FILL; ec.font = GRAY_FONT
        ec.alignment = Alignment(horizontal="center", vertical="center")

        # F 偏差金额
        fc = ws.cell(r, 6, f'=IFERROR(D{r}-B{r},"")')
        fc.number_format = FMT_MONEY; fc.border = BORDER
        fc.fill = GRAY_FILL; fc.font = GRAY_FONT
        fc.alignment = Alignment(horizontal="center", vertical="center")

        # G 偏差%
        gc = ws.cell(r, 7, f'=IFERROR((D{r}-B{r})/B{r},"")')
        gc.number_format = FMT_PCT; gc.border = BORDER
        gc.fill = GRAY_FILL; gc.font = GRAY_FONT
        gc.alignment = Alignment(horizontal="center", vertical="center")

        # H 理论区间备注
        range_note = f"理论区间 {lo:.0%}–{hi:.0%}" if (lo or hi) else ""
        _c(ws, r, 8, range_note, italic=True, fg="595959")

        ws.row_dimensions[r].height = 22
        r += 1

    # 合计行
    end = r - 1
    _c(ws, r, 1, "合 计", bold=True, align="center", bg="DCE6F1", fg="1F3864")
    for ci, (val, fmt, fill) in enumerate([
        (f"=SUM(B4:B{end})",             FMT_MONEY, None),
        (1.0,                             FMT_PCT,   None),
        (f"=SUM(D4:D{end})",             FMT_MONEY, "F2F2F2"),
        (f'=IFERROR(D{r}/D{r},"")',      FMT_PCT,   "F2F2F2"),
        (f'=IFERROR(D{r}-B{r},"")',      FMT_MONEY, "F2F2F2"),
        (f'=IFERROR((D{r}-B{r})/B{r},"")' , FMT_PCT, "F2F2F2"),
    ], start=2):
        c = ws.cell(r, ci, val)
        c.number_format = fmt; c.border = BORDER
        c.font = Font(bold=True, size=10, color="595959" if fill else "1F3864")
        c.fill = _fill(fill if fill else "DCE6F1")
        c.alignment = Alignment(horizontal="center", vertical="center")
    _c(ws, r, 8, "", bg="DCE6F1")
    ws.row_dimensions[r].height = 22
    ws.freeze_panes = "A4"


def build_excel(rows: list[dict], model: str, msrp: float, out_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws_main = wb.active
    ds, de = _build_sheet_main(ws_main, rows, model)
    ws_comp = wb.create_sheet()
    _build_sheet_comparison(ws_comp, rows, model, msrp, ds, de)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


# ══════════════════════════════════════════════════════════════════
#  MSRP 自动查询
# ══════════════════════════════════════════════════════════════════

def _lookup_msrp_from_db(model: str) -> Optional[float]:
    """从 products_db.json 读取零售价"""
    db_path = DATA_DIR / "products_db.json"
    if not db_path.exists():
        return None
    try:
        db = json.loads(db_path.read_text(encoding="utf-8"))
        slug = _slug(model).lower()
        for key, entry in db.items():
            if slug in key.lower() or key.lower() in slug:
                price = entry.get("retail_price_cny")
                if price:
                    return float(price)
    except Exception:
        pass
    return None


def lookup_msrp_from_web(model: str) -> float:
    """用 Claude + web_search 查询零售价，失败返回 5000"""
    print(f"  → 查询 {model} 零售价…")
    try:
        text = _run_web_agent(
            "你是价格查询助手，只输出一个纯数字（人民币元），不要任何其他文字。",
            f"请搜索 {model} 的中国官方建议零售价（CNY），只返回数字。",
            max_tokens=256,
        )
        price = float(re.search(r"\d[\d,\.]*", text.replace(",", "")).group())
        print(f"  ✓ 零售价: {price:.0f} 元")
        return price
    except Exception as e:
        print(f"  ⚠ 价格查询失败: {e}，使用默认值 5000 元")
        return 5000.0


# ══════════════════════════════════════════════════════════════════
#  主流程
# ══════════════════════════════════════════════════════════════════

def main() -> None:
    parser = argparse.ArgumentParser(description="通用拆机分析 Excel 生成器")
    parser.add_argument("model", nargs="?", help="机型名称，如 '石头G30S Pro'")
    parser.add_argument("--msrp", type=float, help="建议零售价（元），不传则自动查询")
    parser.add_argument("--csv", type=Path, help="指定现有 CSV 路径（可选）")
    parser.add_argument("--out", type=Path, help="输出 Excel 路径（默认 data/{slug}_拆机分析.xlsx）")
    parser.add_argument("--enrich", action="store_true", help="强制重新补全所有价格字段")
    parser.add_argument("--no-enrich", dest="no_enrich", action="store_true",
                        help="跳过价格网络补全")
    args = parser.parse_args()

    if not args.model and not args.csv:
        parser.error("请提供机型名称或 --csv 路径")

    model = args.model or args.csv.stem.replace("_teardown", "").replace("_", " ")
    slug  = _slug(model)

    # ── 1. 解析 MSRP ────────────────────────────────────────────
    msrp = args.msrp
    if not msrp:
        msrp = _lookup_msrp_from_db(model)
    if not msrp and not args.no_enrich:
        api_key = os.getenv("ANTHROPIC_API_KEY")
        if api_key:
            msrp = lookup_msrp_from_web(model)
    msrp = msrp or 5000.0

    # ── 2. 解析输出路径 ──────────────────────────────────────────
    out_path = args.out or DATA_DIR / f"{slug}_拆机分析.xlsx"

    print(f"\n机型: {model}  |  零售价: {msrp:.0f} 元  |  输出: {out_path.name}")

    # ── 3. 加载或生成 CSV ────────────────────────────────────────
    csv_path = args.csv
    rows: list[dict] = []
    need_generate = False

    if csv_path and csv_path.exists():
        rows = load_csv(csv_path)
        print(f"  ✓ 加载 CSV: {csv_path.name}（{len(rows)} 条）")
    else:
        csv_path = find_csv(model)
        if csv_path:
            rows = load_csv(csv_path)
            print(f"  ✓ 找到 CSV: {csv_path.name}（{len(rows)} 条）")
        else:
            need_generate = True

    if need_generate:
        api_key = os.getenv("ANTHROPIC_API_KEY")
        if not api_key:
            print("  ✗ 未找到 CSV 且 ANTHROPIC_API_KEY 未设置，无法生成数据")
            sys.exit(1)
        print(f"  → 未找到 {model} 的拆机 CSV，从网络生成…")
        rows = generate_from_web(model, msrp)
        csv_path = TEARDOWN_DIR / f"{slug}_teardown.csv"
        save_csv(rows, csv_path, model)
        print(f"  ✓ 已保存 CSV → {csv_path}")

    # ── 4. 价格补全 ───────────────────────────────────────────────
    if not args.no_enrich:
        api_key = os.getenv("ANTHROPIC_API_KEY")
        if api_key:
            rows = enrich_prices(rows, model, force=args.enrich)
            # 回写更新后的 CSV
            if csv_path:
                save_csv(rows, csv_path, model)
                print(f"  ✓ 已更新 CSV → {csv_path.name}")
        else:
            print("  → ANTHROPIC_API_KEY 未设置，跳过价格补全")

    # ── 5. 生成 Excel ────────────────────────────────────────────
    build_excel(rows, model, msrp, out_path)
    print(f"\n写出 → {out_path}\n")


if __name__ == "__main__":
    main()
