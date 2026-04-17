"""
拆机数据 → 标准件库构建工具

输出目录：
  data/teardowns/  每款机型一个 CSV，文件名含生成日期
  data/lib/        标准件库 components_lib.csv，含 last_updated 列

用法：
  python scripts/build_components.py                    # 处理 data/ 所有拆机 xlsx
  python scripts/build_components.py data/某文件.xlsx   # 处理单文件
"""
from __future__ import annotations

import csv
import os
import sys
from datetime import date
from pathlib import Path
from typing import Any

sys.path.insert(0, str(Path(__file__).parent.parent))
from core.bom_loader import _parse_excel

DATA_DIR     = Path(__file__).parent.parent / "data"
TEARDOWN_DIR = DATA_DIR / "teardowns"
LIB_DIR      = DATA_DIR / "lib"
LIB_CSV      = LIB_DIR / "components_lib.csv"

TODAY = date.today().isoformat()  # e.g. "2026-04-17"

TEARDOWN_FIELDS = [
    "bom_bucket", "section", "name", "model", "type",
    "spec", "manufacturer", "unit_price", "qty", "confidence",
    "product_source", "generated_date",
]

LIB_FIELDS = [
    "id", "bom_bucket", "bom_bucket_cn", "name", "name_en",
    "tier", "model_numbers", "spec", "cost_min", "cost_max", "unit",
    "suppliers", "confidence", "models", "last_updated",
]

# ── BOM 桶映射 ──────────────────────────────────────────────────

BOM_BUCKET_CN = {
    "compute_electronics": "算力与电子",
    "perception":          "感知系统",
    "power_motion":        "动力与驱动",
    "cleaning":            "清洁功能",
    "dock_station":        "基站系统",
    "energy":              "能源系统",
    "structure_cmf":       "整机结构CMF",
    "mva_software":        "MVA+软件授权",
}

# PCB 功能模块 → bom_bucket
PCB_BUCKET_MAP: dict[str, str] = {
    "CPU": "compute_electronics", "SoC": "compute_electronics",
    "NPU": "compute_electronics", "NPU/AI": "compute_electronics",
    "PMIC": "compute_electronics", "DCDC": "compute_electronics",
    "RAM": "compute_electronics", "ROM": "compute_electronics",
    "闪存": "compute_electronics",
    "WIFI": "compute_electronics", "WIFI/BT": "compute_electronics",
    "蓝牙": "compute_electronics", "蓝牙芯片": "compute_electronics",
    "马达驱动": "compute_electronics", "MCU": "compute_electronics",
    "充电IC": "compute_electronics",
    "音频": "compute_electronics", "音频放大": "compute_electronics",
    "音频解码": "compute_electronics",
    "AI语音处理": "perception", "AI语音板": "perception",
    "麦克风": "compute_electronics",
    "IMU": "perception",
    "ToF": "perception", "视觉处理": "perception",
    "多路复用开关": "compute_electronics",
    "运放": "compute_electronics",
    "阻容器件": "compute_electronics",
    "PCB": "compute_electronics",
    "LED": "compute_electronics", "按键": "compute_electronics",
    "红外发送": "perception", "红外发送灯": "perception",
    "LDO": "compute_electronics",
}

# 电机名 → bom_bucket
MOTOR_BUCKET_MAP: dict[str, str] = {
    "风机":         "power_motion",
    "驱动轮电机":   "power_motion",
    "底盘升降电机": "power_motion",
    "边刷电机":     "cleaning",
    "滚刷":         "cleaning",
    "滚刷抬升电机": "cleaning",
    "拖布电机":     "cleaning",
    "拖布震动电机": "cleaning",
    "拖布机械臂电机": "cleaning",
    "拖布抬升电机": "cleaning",
    "拖布伸缩电机": "cleaning",
    "水泵":         "cleaning",
    "气泵":         "cleaning",
}

# 传感器名 → bom_bucket
SENSOR_BUCKET_MAP: dict[str, str] = {
    "雷达":              "perception",
    "前视":              "perception",
    "ToF":               "perception",
    "沿墙":              "perception",
    "驱动轮编码器":      "power_motion",
    "碰撞":              "perception",
    "下视":              "perception",
    "地毯识别":          "perception",
    "拖布安装检测":      "cleaning",
    "滚刷抬起检测":      "cleaning",
    "底盘升降位置检测":  "power_motion",
    "尘盒安装检测":      "perception",
    "回充信号接收":      "perception",
    "基站通讯":          "dock_station",
    "对位检测":          "dock_station",
    "拖布抬起检测":      "cleaning",
    "拖布机械臂到位检测": "cleaning",
    "上盖安装检测":      "perception",
    "边刷位置检测":      "cleaning",
    "超声波驱动板":      "perception",
    "麦克风板":          "compute_electronics",
    "按键显示板":        "compute_electronics",
    "重置/USB/雷达碰撞板": "compute_electronics",
}


def _bucket_pcb(func: str, board: str) -> str:
    f = func.strip()
    if f in PCB_BUCKET_MAP:
        return PCB_BUCKET_MAP[f]
    b = board.strip()
    if "基站" in b:
        return "dock_station"
    if "导航" in b or "视觉" in b:
        return "perception"
    return "compute_electronics"


def _bucket_motor(name: str) -> str:
    for k, v in MOTOR_BUCKET_MAP.items():
        if k in name:
            return v
    return "power_motion"


def _bucket_sensor(name: str) -> str:
    for k, v in SENSOR_BUCKET_MAP.items():
        if k in name:
            return v
    return "perception"


def _bucket_other(name: str) -> str:
    if "电池" in name or "BMS" in name:
        return "energy"
    if "喇叭" in name:
        return "compute_electronics"
    return "structure_cmf"


# ── 拆机数据 → 扁平行列表 ──────────────────────────────────────

def _flatten(model: str, sections: dict) -> list[dict]:
    rows: list[dict] = []

    for item in sections.get("pcb", []):
        func = item.get("function", "") or ""
        if not func:
            continue
        rows.append({
            "bom_bucket":    _bucket_pcb(func, item.get("board", "")),
            "section":       "PCB",
            "name":          func,
            "model":         item.get("model", ""),
            "type":          item.get("sub_board", "") or item.get("board", ""),
            "spec":          item.get("spec", ""),
            "manufacturer":  item.get("manufacturer", ""),
            "unit_price":    item.get("unit_price", ""),
            "qty":           item.get("qty", ""),
            "confidence":    "confirmed" if item.get("unit_price") else "inferred",
            "product_source": model,
        })

    for item in sections.get("motors", []):
        name = item.get("name", "") or ""
        if not name:
            continue
        rows.append({
            "bom_bucket":    _bucket_motor(name),
            "section":       "电机",
            "name":          name,
            "model":         item.get("model", ""),
            "type":          item.get("type", ""),
            "spec":          item.get("params", ""),
            "manufacturer":  item.get("manufacturer", ""),
            "unit_price":    "",
            "qty":           item.get("qty", ""),
            "confidence":    "confirmed",
            "product_source": model,
        })

    for item in sections.get("sensors", []):
        name = item.get("name", "") or ""
        if not name:
            continue
        rows.append({
            "bom_bucket":    _bucket_sensor(name),
            "section":       "传感器",
            "name":          name,
            "model":         "",
            "type":          item.get("type", ""),
            "spec":          item.get("note", ""),
            "manufacturer":  item.get("manufacturer", ""),
            "unit_price":    item.get("unit_price", ""),
            "qty":           item.get("qty", ""),
            "confidence":    "confirmed" if item.get("unit_price") else "inferred",
            "product_source": model,
        })

    for item in sections.get("others", []):
        name = item.get("name", "") or ""
        if not name:
            continue
        rows.append({
            "bom_bucket":    _bucket_other(name),
            "section":       "其他",
            "name":          name,
            "model":         "",
            "type":          item.get("type", ""),
            "spec":          item.get("spec", ""),
            "manufacturer":  item.get("manufacturer", ""),
            "unit_price":    item.get("price", ""),
            "qty":           1,
            "confidence":    "confirmed" if item.get("price") else "inferred",
            "product_source": model,
        })

    return rows


# ── 写 teardown CSV ─────────────────────────────────────────────

def write_teardown_csv(model: str, rows: list[dict], out_dir: Path) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    # 文件名含生成日期，旧版本自动保留（不覆盖）
    path = out_dir / f"{model}_teardown_{TODAY}.csv"
    dated_rows = [{**r, "generated_date": TODAY} for r in rows]
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=TEARDOWN_FIELDS)
        w.writeheader()
        w.writerows(sorted(dated_rows, key=lambda r: (r["bom_bucket"], r["section"], r["name"])))
    return path


# ── 汇总 → components_lib.csv ──────────────────────────────────

def build_lib(all_rows: list[dict]) -> list[dict]:
    registry: dict[str, dict] = {}

    for row in all_rows:
        key = f"{row['bom_bucket']}|{row['section']}|{row['name']}|{row.get('type','')}"
        model = row["product_source"]
        price = row.get("unit_price", "")

        if key not in registry:
            cost_min, cost_max = _parse_price(price)
            registry[key] = {
                "id":           _make_id(row["bom_bucket"], row["name"]),
                "bom_bucket":   row["bom_bucket"],
                "bom_bucket_cn": BOM_BUCKET_CN.get(row["bom_bucket"], row["bom_bucket"]),
                "name":         row["name"],
                "name_en":      "",
                "tier":         "mainstream",
                "model_numbers": row.get("model", ""),
                "spec":         row.get("spec", ""),
                "cost_min":     cost_min,
                "cost_max":     cost_max,
                "unit":         "元/件",
                "suppliers":    row.get("manufacturer", ""),
                "confidence":   row.get("confidence", "inferred"),
                "models":       {model},
            }
        else:
            entry = registry[key]
            entry["models"].add(model)
            # 合并供应商
            if row.get("manufacturer") and row["manufacturer"] not in entry["suppliers"]:
                entry["suppliers"] = "、".join(filter(None, [entry["suppliers"], row["manufacturer"]]))
            # 合并型号
            if row.get("model") and row["model"] not in entry.get("model_numbers", ""):
                entry["model_numbers"] = "、".join(filter(None, [entry.get("model_numbers",""), row["model"]]))
            # 优先用已确认价格
            if not entry["cost_min"] and row.get("unit_price"):
                entry["cost_min"], entry["cost_max"] = _parse_price(row["unit_price"])
            # 置信度：任一机型已确认则升级
            if row.get("confidence") == "confirmed":
                entry["confidence"] = "confirmed"

    lib_rows = []
    for entry in registry.values():
        lib_rows.append({
            **entry,
            "models":       "、".join(sorted(entry["models"])),
            "last_updated": TODAY,
        })

    return sorted(lib_rows, key=lambda r: (r["bom_bucket"], r["name"]))


def _parse_price(price_str) -> tuple[str, str]:
    if not price_str:
        return "", ""
    s = str(price_str).replace("元", "").strip()
    if "~" in s:
        parts = s.split("~")
        return parts[0].strip(), parts[1].strip()
    try:
        v = float(s)
        return str(v), str(v)
    except ValueError:
        return s, ""


def _make_id(bucket: str, name: str) -> str:
    import re
    clean = re.sub(r"[^\w]", "_", name)
    return f"{bucket[:8]}_{clean}"[:48]


# ── 主流程 ──────────────────────────────────────────────────────

def process_xlsx(xlsx_path: Path) -> dict[str, list[dict]]:
    """解析单个 xlsx，返回 {model: [rows]}"""
    orig_env = os.environ.get("BOM_EXCEL_FILE")
    os.environ["BOM_EXCEL_FILE"] = str(xlsx_path)
    data = _parse_excel.__wrapped__(xlsx_path) if hasattr(_parse_excel, "__wrapped__") else None

    # 直接调用解析
    import openpyxl
    from core.bom_loader import _parse_pcb, _parse_motors, _parse_sensors, _parse_others

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    result: dict[str, list[dict]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = [row for row in ws.iter_rows(values_only=True) if any(v is not None for v in row)]

        pcb_rows, motor_rows, sensor_rows, other_rows = [], [], [], []
        current = pcb_rows
        for row in rows:
            first = str(row[0]).strip() if row[0] else ""
            if first in ("电机", "Motor"):   current = motor_rows;  continue
            if first in ("传感器", "Sensor"): current = sensor_rows; continue
            if first in ("其他", "Other"):   current = other_rows;  continue
            current.append(row)

        sections = {
            "pcb":     _parse_pcb(pcb_rows),
            "motors":  _parse_motors(motor_rows),
            "sensors": _parse_sensors(sensor_rows),
            "others":  _parse_others(other_rows),
        }
        flat = _flatten(sheet_name, sections)
        result[sheet_name] = flat

    wb.close()
    if orig_env is not None:
        os.environ["BOM_EXCEL_FILE"] = orig_env
    return result


def main(xlsx_files: list[Path]):
    TEARDOWN_DIR.mkdir(exist_ok=True)
    LIB_DIR.mkdir(exist_ok=True)
    all_rows: list[dict] = []

    try:
        from core.feishu_sync import sync_teardown, sync_components_lib
    except ImportError:
        sync_teardown = sync_components_lib = lambda *a, **kw: None

    for xlsx in xlsx_files:
        print(f"\n处理: {xlsx.name}")
        model_data = process_xlsx(xlsx)
        for model, rows in model_data.items():
            path = write_teardown_csv(model, rows, TEARDOWN_DIR)
            print(f"  ✓ {model}: {len(rows)} 条 → {path.name}")
            all_rows.extend(rows)
            sync_teardown(model, rows)

    lib = build_lib(all_rows)
    with open(LIB_CSV, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=LIB_FIELDS)
        w.writeheader()
        w.writerows(lib)
    print(f"\n标准件库 ({TODAY}): {len(lib)} 条 → {LIB_CSV}")
    sync_components_lib(lib)

    # 统计各桶
    from collections import Counter
    buckets = Counter(r["bom_bucket"] for r in lib)
    for bucket, cnt in sorted(buckets.items()):
        print(f"  {BOM_BUCKET_CN.get(bucket, bucket):12s}  {cnt:3d} 条")


if __name__ == "__main__":
    if len(sys.argv) > 1:
        files = [Path(a) for a in sys.argv[1:]]
    else:
        files = sorted(
            f for f in DATA_DIR.glob("*.xlsx")
            if not f.name.startswith("~$")
        )
        if not files:
            print("data/ 目录下没有找到 xlsx 文件")
            sys.exit(1)

    main(files)
