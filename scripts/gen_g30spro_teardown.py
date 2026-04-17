"""
生成石头 G30S Pro 拆机分析 Excel
格式对齐：机器人价值工程拆机电控分析表（每列含义与 bom_loader.py 保持一致）

数据说明：
  - ✓ 确认：来自官方规格页、权威评测媒体
  - ~ 推测：基于同平台（石头P10Pro/P20）及行业通用方案合理推断
  - ? 待确认：无公开数据，价格为行业区间估算

输出：data/石头G30SPro_拆机分析.xlsx
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = Path(__file__).parent.parent / "data" / "石头G30SPro_拆机分析.xlsx"

# ── 颜色 ──────────────────────────────────────────────────────────
HDR_FILL   = PatternFill("solid", fgColor="4F81BD")
SEC_FILL   = PatternFill("solid", fgColor="DCE6F1")
EST_FILL   = PatternFill("solid", fgColor="FFFFC0")   # 黄：推测值
HDR_FONT   = Font(bold=True, color="FFFFFF")
SEC_FONT   = Font(bold=True)

# ─────────────────────────────────────────────────────────────────
# 数据定义
# 列顺序：[类别1, 功能模块, 芯片型号/名称, 数量, 厂家, 规格参数, 单价/元, 备注]
# ─────────────────────────────────────────────────────────────────

PCB_HEADER = ["PCB", "类别1", "功能模块", "芯片型号", "数量", "厂家", "规格参数", "单价/元", "价格/元"]

# (类别1, 功能模块, 型号, 数量, 厂家, 规格, 单价, 备注)
PCB_DATA = [
    # ── 主板（4层，估算尺寸 220×90mm）────────────────────────────
    ("主板\n4层：220*90mm", "CPU",        "MR813",           1, "全志",    "4核A53，Mali-G31 GPU",          40,  "~同P10Pro平台"),
    ("",                   "NPU/AI",     "BM1688/自研",      1, "比特大陆/自研", "5TOPS+，用于280+物体识别",    60,  "~旗舰AI算力升级，型号待确认"),
    ("",                   "PMIC",       "AXP305B",          1, "全志",    "",                               4,   "~"),
    ("",                   "RAM",        "GDQ2BFAA",         2, "兆易创新", "DDR4、4Gb",                     8,   "~同P10Pro"),
    ("",                   "ROM",        "KLM8G1GETF-B041",  1, "三星",    "EMMC5.1、8GB",                  17,  "~"),
    ("",                   "IMU",        "ICM-42688-P",      1, "InvenSense", "6轴，陀螺仪+加速度计",         12,  "~旗舰机常用方案"),
    ("",                   "DCDC",       "",                 3, "",        "多路电源转换",                    6,   "~"),
    ("",                   "WIFI/BT",    "RTL8821CS",        1, "Realtek", "2.4G/5G双频 Wi-Fi 5, BT 5.0",   8,   "~"),
    ("",                   "马达驱动",   "DRV8833",          2, "TI",      "双通道H桥，2A",                  6,   "~"),
    ("",                   "充电IC",     "BQ25895",          1, "TI",      "快充，5A",                       10,  "~"),
    ("",                   "MCU",        "STM32G0B1",        1, "ST",      "Cortex-M0+，电机/传感器控制",    8,   "~"),
    ("",                   "音频",       "ES8388",           1, "Everest", "音频编解码",                     4,   "~"),
    ("",                   "麦克风",     "MSM421A",          2, "",        "MEMS数字麦克风",                  3,   "~"),
    ("",                   "阻容器件",   "",                 1, "",        "电阻/电容/晶振",                  8,   "~"),
    ("",                   "PCB",        "",                 1, "",        "4层板",                          18,  "~"),
    # ── 导航板 ────────────────────────────────────────────────────
    ("导航/视觉板",        "视觉处理",   "RV1109",           1, "瑞芯微",  "双核A7+NNIE，用于结构光处理",    35,  "~或集成在主SoC"),
    ("",                   "ToF",        "VL53L5CX",         1, "ST",      "顶部ToF，8×8 zone",              15,  "~"),
    ("",                   "PCB",        "",                 1, "",        "导航子板",                        8,  "~"),
    # ── 按键显示板 ─────────────────────────────────────────────────
    ("按键显示板",         "MCU",        "STM32F030",        1, "ST",      "按键+LED控制",                   3,   "~"),
    ("",                   "LED",        "",                 3, "",        "状态指示灯",                      1,   "~"),
    ("",                   "按键",       "",                 2, "",        "触控/机械按键",                   1,   "~"),
    # ── 基站通讯板 ─────────────────────────────────────────────────
    ("基站通讯板",         "红外发送",   "",                 2, "",        "红外对准/回充信号发射",            1,   "✓"),
]

MOTOR_HEADER = ["电机", None, "电机类型", "型号", "数量", "厂家", "规格参数", "单价/元", "价格/元"]

# (名称, 类型, 型号, 规格, 数量, 厂家, 单价, 备注)
MOTOR_DATA = [
    ("边刷电机",     "直流有刷",   "",  "空载转速≥300rpm",     1, "建准/万宝至", 8,   "✓类型 ~价格"),
    ("驱动轮电机",   "直流无刷",   "",  "扭矩≥0.3N·m，编码器", 2, "",            22,  "~旗舰无刷方案"),
    ("风机",         "直流无刷BLDC","", "35000Pa，高压风机",    1, "建准",        55,  "✓吸力 ~供应商"),
    ("滚刷电机",     "直流有刷",   "",  "",                     1, "",            10,  "~"),
    ("拖布震动电机", "直流无刷",   "",  "4000次/min，双震动区", 2, "",            12,  "✓参数"),
    ("水泵",         "直流有刷蠕动泵","","",                    1, "",            8,   "~"),
    ("底盘升降电机", "直流有刷",   "",  "升降行程≥45mm",        1, "",            15,  "✓功能 ~电机类型"),
    ("拖布伸缩电机", "步进电机",   "",  "侧向伸出，边角清洁",   1, "",            12,  "✓功能 ~类型"),
]

SENSOR_HEADER = ["传感器", None, "传感器类型", "数量", "厂家", "单价/元", "备注"]

# (名称, 类型, 数量, 厂家, 单价, 备注)
SENSOR_DATA = [
    ("雷达",              "升降式激光雷达",         1, "自研",     120, "✓升降LDS"),
    ("前视",              "三线结构光+RGB+补光灯",  1, "",          45, "✓三线结构光方案"),
    ("ToF",               "顶部ToF（多区）",         1, "ST",        15, "~VL53L5CX"),
    ("沿墙",              "斜向激光",               1, "",           8, "✓斜向激光方案"),
    ("驱动轮编码器",      "霍尔",                   2, "",           2, "~"),
    ("碰撞",              "微动开关",               2, "",           1, "~"),
    ("下视",              "红外对管",               4, "",           1, "~"),
    ("地毯识别",          "超声波",                 2, "",           3, "~"),
    ("拖布安装检测",      "霍尔",                   1, "",           1, "~"),
    ("滚刷抬起检测",      "光耦",                   1, "",           1, "~"),
    ("底盘升降位置检测",  "霍尔",                   1, "",           1, "~"),
    ("尘盒安装检测",      "霍尔",                   1, "",           1, "~"),
    ("回充信号接收",      "红外探头",               2, "",           1, "✓"),
    ("基站通讯",          "红外",                   2, "",           1, "✓"),
]

OTHER_HEADER = ["其他", None, "类型", "规格", "数量", "厂家", "单价/元", "备注"]

# (名称, 类型, 规格, 数量, 厂家, 单价, 备注)
OTHER_DATA = [
    ("电池包", "18650、4串2并", "额定容量6400mAh、14.4V，快充", 1, "电芯：宁德/比克", 165, "✓6400mAh"),
    ("喇叭",   "直流",          "8Ω、1W",                        1, "",                4,   "~"),
]


# ─────────────────────────────────────────────────────────────────
def _write_section(ws, start_row: int, section_label: str,
                   header: list, data: list, col_count: int) -> int:
    """写一个分段，返回下一个可用行号"""
    row = start_row

    # 分段标记（A列合并）
    ws.cell(row, 1, section_label)
    # 写表头（从B列开始，跳过A）
    for ci, val in enumerate(header[1:], 2):
        c = ws.cell(row, ci, val)
        if val:
            c.font = HDR_FONT
            c.fill = HDR_FILL
            c.alignment = Alignment(horizontal="center")
    ws.cell(row, 1).font = SEC_FONT
    ws.cell(row, 1).fill = SEC_FILL

    row += 1
    section_start = row

    for item in data:
        # col B onward
        note = item[-1] if len(item) > 6 else ""
        is_estimated = "~" in str(note)

        for ci, val in enumerate(item[:-1], 2):  # skip last (备注 in data)
            c = ws.cell(row, ci, val)
            if is_estimated:
                c.fill = EST_FILL
            c.alignment = Alignment(wrap_text=True, vertical="center")
        # 备注写在最后一列
        note_col = 2 + len(item) - 1
        ws.cell(row, note_col, note)
        row += 1

    # 合并A列（分段标记）
    if row - section_start > 0:
        ws.merge_cells(
            start_row=start_row, start_column=1,
            end_row=row - 1, end_column=1
        )
    ws.cell(start_row, 1).alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )

    # 空白分隔行
    row += 1
    return row


def generate():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "石头G30SPro"

    # 列宽
    widths = [8, 22, 18, 18, 6, 14, 28, 10, 10, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 说明行
    ws.cell(1, 1, "石头G30S Pro 拆机电控分析（基于公开规格+同平台推测）")
    ws.cell(1, 1).font = Font(bold=True, size=12)
    ws.cell(2, 1, "✓=公开确认  ~=同平台推测  ?=待确认  黄色背景=推测值")
    ws.cell(2, 1).font = Font(italic=True, color="666666")
    ws.merge_cells("A1:J1")
    ws.merge_cells("A2:J2")

    row = 4
    row = _write_section(ws, row, "PCB",  PCB_HEADER,    PCB_DATA,    9)
    row = _write_section(ws, row, "电机", MOTOR_HEADER,  MOTOR_DATA,  9)
    row = _write_section(ws, row, "传感器", SENSOR_HEADER, SENSOR_DATA, 7)
    row = _write_section(ws, row, "其他", OTHER_HEADER,  OTHER_DATA,  8)

    # 冻结前两行
    ws.freeze_panes = "B4"

    OUT.parent.mkdir(exist_ok=True)
    wb.save(OUT)
    print(f"写出 → {OUT}")
    return OUT


if __name__ == "__main__":
    generate()
