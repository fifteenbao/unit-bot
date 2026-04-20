"""
从产品数据库 CSV（或 xlsx）批量导入产品到 products_db.json

CSV 格式：GBK 编码，首行为列名，每行一款 SKU。
xlsx 格式：首 Sheet 首行列名，结构相同（兼容旧格式）。

用法：
    python scripts/import_products.py data/产品数据库.csv
    python scripts/import_products.py data/产品数据库.xlsx
    python scripts/import_products.py          # 自动查找 data/*.csv 或 data/*.xlsx
"""
from __future__ import annotations

import csv
import re
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))
from core.db import upsert_product, update_completeness


# ── 辅助函数 ──────────────────────────────────────────────────────

def _bool(val) -> bool | None:
    if val is None:
        return None
    s = str(val).strip()
    if s == "是":
        return True
    if s == "否":
        return False
    return None


def _int_from_str(val, pattern: str) -> int | None:
    if val is None:
        return None
    m = re.search(pattern, str(val))
    return int(float(m.group(1))) if m else None


def _slug(brand: str, name: str) -> str:
    clean = re.sub(r"[^\w\u4e00-\u9fff]", "", name)
    return f"{brand}_{clean}"[:64]


def _parse_row(d: dict) -> dict | None:
    name  = (d.get("产品名称") or "").strip()
    brand = (d.get("厂商名称") or "").strip()
    if not name or not brand:
        return None

    suction     = _int_from_str(d.get("吸力"), r"(\d+)")
    battery_mah = _int_from_str(d.get("电池容量"), r"(\d+)")
    life        = _int_from_str(d.get("续航"), r"(\d+)")
    obs         = _int_from_str(d.get("越障高度"), r"(\d+(?:\.\d+)?)")

    release = d.get("发布日期") or d.get("SKU 更新时间")
    if isinstance(release, datetime):
        release_str = release.strftime("%Y-%m")
    elif release:
        release_str = str(release).strip()[:7]
    else:
        release_str = None

    url_raw = str(d.get("商品链接") or d.get("产品链接") or "")
    url_m   = re.search(r"https?://[^\s\"]+", url_raw)
    url     = url_m.group(0).rstrip('"') if url_m else None

    specs = {
        "suction_power_pa":     suction,
        "battery_capacity_mah": battery_mah,
        "battery_life_min":     life,
        "obstacle_height_cm":   obs,
        "navigation":           (d.get("导航方式") or "").strip() or None,
        "mop_lift":             _bool(d.get("拖布抬升")),
        "self_cleaning":        _bool(d.get("是否自清洁")),
        "hot_air_dry":          _bool(d.get("自动烘干")),
        "auto_empty":           _bool(d.get("自动集尘")),
        "auto_wash":            _bool(d.get("自动清洗拖布")),
    }
    specs = {k: v for k, v in specs.items() if v is not None}

    features = {
        "自动上下水":       _bool(d.get("自动上下水")),
        "自动添加清洁液":   _bool(d.get("自动添加清洁液")),
        "自动补水":         _bool(d.get("自动补水")),
        "底盘升降":         _bool(d.get("底盘升降")),
        "热水擦地":         _bool(d.get("热水擦地")),
        "热风烘干":         _bool(d.get("热风烘干")),
        "边角清洁":         _bool(d.get("边角清洁")),
        "毛发防缠":         _bool(d.get("毛发防缠")),
        "地毯加压清扫":     _bool(d.get("地毯加压清扫")),
        "智能避障":         _bool(d.get("智能避障")),
        "物体识别":         _bool(d.get("物体识别")),
        "语音交互":         _bool(d.get("语音交互")),
        "高温自清洁基站":   _bool(d.get("高温自清洁基站")),
    }
    features = {k: v for k, v in features.items() if v is not None}

    price_raw = d.get("价格")
    try:
        price = float(str(price_raw).replace("元", "").strip()) if price_raw else None
    except ValueError:
        price = None

    return {
        "key": _slug(brand, name),
        "data": {
            "brand":            brand,
            "model_name":       name,
            "retail_price_cny": price,
            "release_date":     release_str,
            "market_segment":   None,
            "product_page_url": url,
            "data_sources": {
                "web_research": [url] if url else [],
                "completeness": {},
            },
            "specs":    specs,
            "features": features,
            "notes":    str(d.get("卖点摘要") or "").strip(),
        },
    }


# ── 读取器 ────────────────────────────────────────────────────────

def _rows_from_csv(path: Path):
    for enc in ("utf-8-sig", "gbk", "gb18030"):
        try:
            with path.open(encoding=enc) as f:
                reader = csv.DictReader(f)
                rows = list(reader)
            return rows
        except (UnicodeDecodeError, Exception):
            continue
    raise ValueError(f"无法解码 {path}，请确认编码为 UTF-8 或 GBK")


def _rows_from_xlsx(path: Path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(row):
            rows.append(dict(zip(headers, row)))
    wb.close()
    return rows


# ── 主流程 ────────────────────────────────────────────────────────

def import_file(path: Path) -> int:
    if path.suffix.lower() == ".csv":
        rows = _rows_from_csv(path)
    elif path.suffix.lower() in (".xlsx", ".xls"):
        rows = _rows_from_xlsx(path)
    else:
        print(f"不支持的文件格式：{path.suffix}")
        return 0

    imported = 0
    for row in rows:
        parsed = _parse_row(row)
        if not parsed:
            continue
        key = parsed["key"]
        upsert_product(key, parsed["data"])
        update_completeness(key)
        imported += 1
        print(f"  ✓ {key}")

    return imported


if __name__ == "__main__":
    if len(sys.argv) > 1:
        path = Path(sys.argv[1])
    else:
        data_dir = Path(__file__).parent.parent / "data"
        candidates = (
            sorted(data_dir.glob("产品数据库.csv"))
            or sorted(f for f in data_dir.glob("*.csv") if "产品" in f.name)
            or sorted(f for f in data_dir.glob("*.xlsx") if "产品" in f.name)
        )
        path = candidates[0] if candidates else None

    if not path or not path.exists():
        print("未找到产品数据库文件，请传入路径：")
        print("  python scripts/import_products.py data/产品数据库.csv")
        sys.exit(1)

    print(f"导入来源：{path}")
    count = import_file(path)
    print(f"\n完成：共导入 {count} 条产品")
